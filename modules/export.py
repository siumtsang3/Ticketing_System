import os
import datetime

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QMessageBox, QFrame, QDateEdit, QFileDialog,
    QGroupBox, QRadioButton, QButtonGroup,
)
from PyQt6.QtCore import QDate

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Colour palette for the worksheet ─────────────────────────────────────────
HDR_BG   = "2D3748"
HDR_FG   = "FFFFFF"
ROW_A_BG = "EBF4FF"   # even rows – light blue
ROW_B_BG = "FFFFFF"   # odd rows  – white
BORDER_COLOR = "CBD5E0"


def _thin_border():
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _write_excel(rows: list[dict], title: str, filepath: str):
    """Build and save an Excel workbook from a list of task dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # ── Column definitions (header text, field key, width) ────────────
    cols = [
        ("Task ID",        "task_id",          8),
        ("Title",          "task_title",       40),
        ("Job Type",       "job_type",         22),
        ("Priority",       "priority",         12),
        ("Status",         "status",           18),
        ("Requested By",   "req_by_name",      22),
        ("Followers",      "fup_by_name",      28),
        ("Requested Date", "requested_date",   16),
        ("Due Date",       "due_date",         16),
        ("Project",        "project_name",     30),
        ("Description",    "task_description", 50),
        ("Progress",       "progress_log",     60),
        ("Remarks",        "remarks",          40),
    ]

    # ── Report title row ─────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=13, color=HDR_FG)
    title_cell.fill      = PatternFill("solid", fgColor=HDR_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Generated timestamp
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ts_cell = ws.cell(
        row=2, column=1,
        value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ts_cell.font      = Font(italic=True, size=10, color="718096")
    ts_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Header row ───────────────────────────────────────────────────
    HDR_ROW = 3
    for c_idx, (hdr, _, col_w) in enumerate(cols, start=1):
        cell = ws.cell(row=HDR_ROW, column=c_idx, value=hdr)
        cell.font      = Font(bold=True, color=HDR_FG, size=11)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(c_idx)].width = col_w
    ws.row_dimensions[HDR_ROW].height = 22

    # ── Data rows ────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=HDR_ROW + 1):
        bg = ROW_A_BG if r_idx % 2 == 0 else ROW_B_BG
        fill = PatternFill("solid", fgColor=bg)
        for c_idx, (_, key, _) in enumerate(cols, start=1):
            val = row.get(key)
            if val is None:
                val = ""
            elif hasattr(val, "strftime"):
                val = str(val)[:10]
            cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val != "" else "")
            cell.fill      = fill
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Freeze header rows ───────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Auto-filter ──────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{HDR_ROW}:{get_column_letter(len(cols))}{HDR_ROW}"
    )

    wb.save(filepath)


# ──────────────────────────────────────────────────────────────────────────────
class ExportModule(QWidget):
    """Excel export – by date range or by project."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._build_ui()
        self._refresh_projects()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        title = QLabel("Export to Excel")
        title.setObjectName("pageTitle")
        root.addWidget(title)

        line = QFrame(); line.setObjectName("hLine"); line.setFrameShape(QFrame.Shape.HLine)
        root.addWidget(line)

        if not OPENPYXL_OK:
            warn = QLabel("⚠  openpyxl is not installed.\nRun:  pip install openpyxl")
            warn.setStyleSheet("color: #c53030; font-size: 14px;")
            root.addWidget(warn)
            root.addStretch()
            return

        # ── Mode selection ────────────────────────────────────────────
        mode_group = QGroupBox("Export Mode")
        mode_lay = QHBoxLayout(mode_group)
        self._rb_date    = QRadioButton("By Date Range")
        self._rb_project = QRadioButton("By Project")
        self._rb_date.setChecked(True)
        bg = QButtonGroup(self)
        bg.addButton(self._rb_date)
        bg.addButton(self._rb_project)
        self._rb_date.toggled.connect(self._on_mode_change)
        mode_lay.addWidget(self._rb_date)
        mode_lay.addWidget(self._rb_project)
        mode_lay.addStretch()
        root.addWidget(mode_group)

        # ── Date range section ────────────────────────────────────────
        self._grp_date = QGroupBox("Date Range  (based on Requested Date)")
        date_form = QHBoxLayout(self._grp_date)
        date_form.setSpacing(12)

        self.de_from = QDateEdit()
        self.de_to   = QDateEdit()
        for de in (self.de_from, self.de_to):
            de.setCalendarPopup(True)
            de.setDisplayFormat("dd/MM/yyyy")
        # Defaults: first day of current month → today
        today = QDate.currentDate()
        self.de_from.setDate(QDate(today.year(), today.month(), 1))
        self.de_to.setDate(today)

        date_form.addWidget(QLabel("From:"))
        date_form.addWidget(self.de_from)
        date_form.addWidget(QLabel("To:"))
        date_form.addWidget(self.de_to)
        date_form.addStretch()
        root.addWidget(self._grp_date)

        # ── Project section ───────────────────────────────────────────
        self._grp_project = QGroupBox("Project")
        proj_lay = QHBoxLayout(self._grp_project)
        proj_lay.setSpacing(12)
        self.cmb_project = QComboBox()
        self.cmb_project.setMinimumWidth(280)
        proj_lay.addWidget(QLabel("Select Project:"))
        proj_lay.addWidget(self.cmb_project)
        proj_lay.addStretch()
        self._grp_project.setEnabled(False)
        root.addWidget(self._grp_project)

        # ── Export button ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self.btn_export = QPushButton("Export to Excel (.xlsx)")
        self.btn_export.setObjectName("successButton")
        self.btn_export.setFixedWidth(240)
        self.btn_export.clicked.connect(self._export)
        btn_row.addWidget(self.btn_export)
        btn_row.addStretch()
        root.addLayout(btn_row)

        self.lbl_status = QLabel("")
        self.lbl_status.setObjectName("footerLabel")
        root.addWidget(self.lbl_status)

        root.addStretch()

    # ── Helpers ───────────────────────────────────────────────────────
    def _on_mode_change(self):
        by_date = self._rb_date.isChecked()
        self._grp_date.setEnabled(by_date)
        self._grp_project.setEnabled(not by_date)

    def _refresh_projects(self):
        if not OPENPYXL_OK:
            return
        self.cmb_project.clear()
        try:
            for p in self.db.get_projects_list():
                self.cmb_project.addItem(p["project_name"], p["project_id"])
        except Exception:
            pass

    # ── Export ────────────────────────────────────────────────────────
    def _export(self):
        self.lbl_status.setText("")
        try:
            rows, report_title = self._fetch_data()
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc)); return

        if not rows:
            QMessageBox.information(self, "No Data",
                                    "No tasks found for the selected criteria.")
            return

        # Ask user for save path
        default_name = (
            f"Tasks_Export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        )
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File",
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)",
        )
        if not filepath:
            return
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"

        try:
            _write_excel(rows, report_title, filepath)
            self.lbl_status.setText(
                f"Exported {len(rows)} task(s) → {filepath}"
            )
            reply = QMessageBox.question(
                self, "Export Complete",
                f"Exported {len(rows)} task(s).\n\nOpen the file now?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    os.startfile(filepath)
                except Exception:
                    pass
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))

    def _fetch_data(self):
        # Long progress logs can exceed the default GROUP_CONCAT cap (1024 bytes)
        try:
            self.db.execute("SET SESSION group_concat_max_len = 1000000")
        except Exception:
            pass

        base_q = (
            "SELECT t.*, "
            "p1.full_name AS req_by_name, "
            "(SELECT GROUP_CONCAT(pe.full_name ORDER BY pe.full_name SEPARATOR ', ') "
            " FROM task_followers tf JOIN people pe ON pe.person_id = tf.person_id "
            " WHERE tf.task_id = t.task_id) AS fup_by_name, "
            "COALESCE("
            " (SELECT GROUP_CONCAT("
            "    CONCAT('[', u.update_date, '] ', u.update_details) "
            "    ORDER BY u.update_date, u.update_id SEPARATOR '\n') "
            "  FROM task_updates u WHERE u.task_id = t.task_id), "
            " t.current_progress) AS progress_log, "
            "pr.project_name "
            "FROM tasks t "
            "LEFT JOIN people   p1 ON t.requested_by   = p1.person_id "
            "LEFT JOIN projects pr ON t.linked_project = pr.project_id "
        )

        if self._rb_date.isChecked():
            d_from = self.de_from.date().toString("yyyy-MM-dd")
            d_to   = self.de_to.date().toString("yyyy-MM-dd")
            if d_from > d_to:
                raise ValueError("'From' date must be on or before 'To' date.")
            q      = base_q + "WHERE t.requested_date BETWEEN %s AND %s ORDER BY t.requested_date"
            params = (d_from, d_to)
            title  = f"Task Export  |  {d_from}  to  {d_to}"
        else:
            proj_id   = self.cmb_project.currentData()
            proj_name = self.cmb_project.currentText()
            if proj_id is None:
                raise ValueError("No project selected.")
            q      = base_q + "WHERE t.linked_project=%s ORDER BY t.requested_date"
            params = (proj_id,)
            title  = f"Task Export  |  Project: {proj_name}"

        rows = self.db.execute(q, params, fetch=True)
        return rows, title
