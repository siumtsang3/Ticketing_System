"""
Mass-import tasks from an Excel (.xlsx) file, plus a template generator.

Public API
----------
generate_template(db, filepath)      -> writes a blank import template
import_tasks(db, filepath)           -> (imported_count, messages)
ImportTasksDialog(db, parent)        -> QDialog UI (template + import)
"""
import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTextEdit, QFileDialog, QMessageBox, QFrame,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from modules.tasks import JOB_TYPES, PRIORITIES, STATUSES


# (header text, column width) – column order in the template
IMPORT_HEADERS = [
    ("Task Title*",    34),
    ("Description",    40),
    ("Job Type",       22),
    ("Priority",       12),
    ("Status",         18),
    ("Requested By",   20),
    ("Followers",      28),
    ("Requested Date", 16),
    ("Due Date",       16),
    ("Linked Project", 26),
    ("Remarks",        30),
]

_MAX_ROWS = 200   # rows the dropdowns / validation cover


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def _norm(header) -> str:
    """Normalise a header cell to a lookup key (strip, drop '*', lowercase)."""
    return (str(header) if header is not None else "").strip().rstrip("*").strip().lower()


def _parse_date(val):
    """Return 'YYYY-MM-DD' for a date/datetime or common string formats, else None."""
    if val is None or val == "":
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Template generation
# ──────────────────────────────────────────────────────────────────────────────
def generate_template(db, filepath):
    wb = openpyxl.Workbook()

    # --- Tasks sheet (data entry) ---------------------------------------
    ws = wb.active
    ws.title = "Tasks"

    hdr_fill = PatternFill("solid", fgColor="2D3748")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CBD5E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (text, width) in enumerate(IMPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=text)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # --- Instructions sheet ---------------------------------------------
    ins = wb.create_sheet("Instructions")
    lines = [
        ("IT Ticketing System — Task Import Template", True),
        ("", False),
        ("• Enter one task per row on the 'Tasks' sheet.", False),
        ("• Only 'Task Title' is required — every other column is optional.", False),
        ("• Coloured dropdowns are provided for Job Type, Priority, Status,", False),
        ("  Requested By and Linked Project.", False),
        ("", False),
        ("Job Type        : " + " / ".join(JOB_TYPES), False),
        ("Priority        : " + " / ".join(PRIORITIES) + "    (blank = Medium)", False),
        ("Status          : " + " / ".join(STATUSES) + "    (blank = Open)", False),
        ("Requested By     : a person's full name (must exist in People)", False),
        ("Followers        : one or more names separated by comma or semicolon", False),
        ("Linked Project   : a project name (must exist in Projects)", False),
        ("Requested / Due  : date as YYYY-MM-DD or DD/MM/YYYY", False),
        ("                   (blank Requested Date defaults to today)", False),
        ("", False),
        ("Names that do not match existing People / Projects are imported", False),
        ("with the field left blank and a note in the import summary.", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        c = ins.cell(row=r, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13)
    ins.column_dimensions["A"].width = 95

    # --- Lists sheet (dropdown sources, hidden) -------------------------
    lists = wb.create_sheet("Lists")
    lists["A1"] = "People"
    people = [p["full_name"] for p in db.get_people_list()]
    for r, name in enumerate(people, start=2):
        lists.cell(row=r, column=1, value=name)
    lists["B1"] = "Projects"
    projects = [p["project_name"] for p in db.get_projects_list()]
    for r, name in enumerate(projects, start=2):
        lists.cell(row=r, column=2, value=name)
    lists.sheet_state = "hidden"

    # --- Data validation dropdowns --------------------------------------
    def add_list_dv(col_letter, formula1):
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.showErrorMessage = True
        dv.error = "Please pick a value from the dropdown list."
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{_MAX_ROWS}")

    add_list_dv("C", '"%s"' % ",".join(JOB_TYPES))    # Job Type
    add_list_dv("D", '"%s"' % ",".join(PRIORITIES))   # Priority
    add_list_dv("E", '"%s"' % ",".join(STATUSES))     # Status
    if people:
        add_list_dv("F", f"=Lists!$A$2:$A${1 + len(people)}")    # Requested By
    if projects:
        add_list_dv("J", f"=Lists!$B$2:$B${1 + len(projects)}")  # Linked Project

    # --- Sheet order: Instructions first, land the user on Tasks --------
    wb.move_sheet("Instructions", -1)
    wb.active = wb.sheetnames.index("Tasks")

    wb.save(filepath)


# ──────────────────────────────────────────────────────────────────────────────
# Import
# ──────────────────────────────────────────────────────────────────────────────
def import_tasks(db, filepath):
    """Read tasks from filepath. Returns (imported_count, list_of_messages)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return 0, ["The worksheet is empty."]

    idx = {}
    for i, h in enumerate(rows[0]):
        key = _norm(h)
        if key:
            idx[key] = i
    if "task title" not in idx:
        return 0, ["Could not find a 'Task Title' column in the header row."]

    people = {p["full_name"].strip().lower(): p["person_id"]
              for p in db.get_people_list()}
    projects = {p["project_name"].strip().lower(): p["project_id"]
                for p in db.get_projects_list()}

    def getval(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return None
        v = row[i]
        if isinstance(v, str):
            v = v.strip()
        return v if v not in ("", None) else None

    imported = 0
    messages = []
    for rno, row in enumerate(rows[1:], start=2):
        if row is None or all(
            c is None or (isinstance(c, str) and not c.strip()) for c in row
        ):
            continue  # entirely blank row

        title = getval(row, "task title")
        if not title:
            messages.append(f"Row {rno}: skipped — no Task Title.")
            continue

        warns = []
        desc = getval(row, "description") or ""

        job = getval(row, "job type") or ""
        if job and job not in JOB_TYPES:
            warns.append(f"unknown Job Type '{job}' (kept as typed)")

        prio = getval(row, "priority") or "Medium"
        if prio not in PRIORITIES:
            warns.append(f"unknown Priority '{prio}', used Medium")
            prio = "Medium"

        status = getval(row, "status") or "Open"
        if status not in STATUSES:
            warns.append(f"unknown Status '{status}', used Open")
            status = "Open"

        rb_name = getval(row, "requested by")
        rb_id = None
        if rb_name:
            rb_id = people.get(str(rb_name).lower())
            if rb_id is None:
                warns.append(f"Requested By '{rb_name}' not found")

        foll_ids = []
        foll_raw = getval(row, "followers")
        if foll_raw:
            parts = [x.strip() for x in str(foll_raw).replace(";", ",").split(",")
                     if x.strip()]
            for nm in parts:
                pid = people.get(nm.lower())
                if pid:
                    if pid not in foll_ids:
                        foll_ids.append(pid)
                else:
                    warns.append(f"Follower '{nm}' not found")

        rd_raw = getval(row, "requested date")
        rd = _parse_date(rd_raw)
        if rd is None and rd_raw is not None:
            warns.append(f"unreadable Requested Date '{rd_raw}', used today")
        if rd is None:
            rd = datetime.date.today().strftime("%Y-%m-%d")

        dd_raw = getval(row, "due date")
        dd = _parse_date(dd_raw)
        if dd is None and dd_raw is not None:
            warns.append(f"unreadable Due Date '{dd_raw}', left blank")

        pj_name = getval(row, "linked project")
        pj_id = None
        if pj_name:
            pj_id = projects.get(str(pj_name).lower())
            if pj_id is None:
                warns.append(f"Project '{pj_name}' not found")

        remarks = getval(row, "remarks") or ""
        primary = foll_ids[0] if foll_ids else None

        try:
            tid = db.execute(
                "INSERT INTO tasks (task_title, task_description, job_type, priority, "
                "status, requested_by, follow_up_by, requested_date, due_date, "
                "linked_project, remarks) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (title, desc, job, prio, status, rb_id, primary, rd, dd, pj_id, remarks),
            )
            if foll_ids:
                db.set_task_followers(tid, foll_ids)
            imported += 1
            if warns:
                messages.append(
                    f"Row {rno}: imported as task #{tid} — note: " + "; ".join(warns)
                )
        except Exception as exc:
            messages.append(f"Row {rno}: FAILED — {exc}")

    return imported, messages


# ──────────────────────────────────────────────────────────────────────────────
# Dialog
# ──────────────────────────────────────────────────────────────────────────────
class ImportTasksDialog(QDialog):
    """Download a template, then import a filled-in Excel file."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.imported_any = False
        self.setWindowTitle("Import Tasks from Excel")
        self.setMinimumWidth(620)
        self.setMinimumHeight(460)
        self.setModal(True)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 20, 24, 16)
        lay.setSpacing(12)

        title = QLabel("Mass-import Tasks")
        title.setObjectName("pageTitle")
        lay.addWidget(title)

        info = QLabel(
            "Step 1 — Save the blank template and fill in your tasks "
            "(one per row).\n"
            "Step 2 — Select the completed file to import it.\n"
            "Only 'Task Title' is required; names are matched to existing "
            "People / Projects."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color: #4a5568;")
        lay.addWidget(info)

        btn_row = QHBoxLayout()
        btn_tpl = QPushButton("1.  Save Blank Template…")
        btn_tpl.setObjectName("secondaryButton")
        btn_imp = QPushButton("2.  Select Excel & Import…")
        btn_imp.setObjectName("successButton")
        btn_tpl.clicked.connect(self._save_template)
        btn_imp.clicked.connect(self._do_import)
        btn_row.addWidget(btn_tpl)
        btn_row.addWidget(btn_imp)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        line = QFrame(); line.setObjectName("hLine"); line.setFrameShape(QFrame.Shape.HLine)
        lay.addWidget(line)

        lay.addWidget(QLabel("Result:"))
        self.txt_result = QTextEdit()
        self.txt_result.setReadOnly(True)
        lay.addWidget(self.txt_result)

        close_row = QHBoxLayout()
        close_row.addStretch()
        btn_close = QPushButton("Close")
        btn_close.setObjectName("secondaryButton")
        btn_close.clicked.connect(self.accept)
        close_row.addWidget(btn_close)
        lay.addLayout(close_row)

    # ── actions ─────────────────────────────────────────────────────────
    def _save_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Task Import Template",
            "Task_Import_Template.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            generate_template(self.db, path)
            self.txt_result.append(f"✓ Template saved:\n   {path}\n")
            QMessageBox.information(self, "Template Saved",
                                    f"Template saved to:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Could not save template:\n{exc}")

    def _do_import(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File to Import",
            "", "Excel Files (*.xlsx *.xlsm)"
        )
        if not path:
            return
        try:
            count, messages = import_tasks(self.db, path)
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            return

        self.txt_result.append(f"✓ Imported {count} task(s) from:\n   {path}")
        if messages:
            self.txt_result.append("\nNotes:")
            for m in messages:
                self.txt_result.append("  • " + m)
        self.txt_result.append("")
        if count:
            self.imported_any = True
        QMessageBox.information(
            self, "Import Complete",
            f"Imported {count} task(s)."
            + (f"\n{len(messages)} row(s) had notes — see the Result panel."
               if messages else "")
        )
